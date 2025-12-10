from fastapi import FastAPI, HTTPException, WebSocket
from fastapi.responses import JSONResponse, FileResponse
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.responses import Response
from pydantic import BaseModel
from typing import List, Set, Optional
import uuid
from fastapi.middleware.cors import CORSMiddleware
from datetime import datetime, timedelta
import math
import os
import pandas as pd
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
import io
import json
import threading
import paho.mqtt.client as mqtt
import smtplib
from email.mime.text import MIMEText
from passlib.context import CryptContext
import ssl

pwd_context = CryptContext(schemes=["pbkdf2_sha256"], deprecated="auto")

# Pydantic models for API
class AuthRequest(BaseModel):
    username: str
    password: str

class AuthResponse(BaseModel):
    access_token: str
    refresh_token: str
    username: str
    role: str

class User(BaseModel):
    id: int
    username: str
    email: Optional[str] = None
    role: str = "admin"
    is_active: bool = True
    created_at: datetime

class UserCreate(BaseModel):
    username: str
    password: str
    email: Optional[str] = None
    role: str = "viewer"

class UserUpdate(BaseModel):
    email: Optional[str] = None
    role: Optional[str] = None
    is_active: Optional[bool] = None
    password: Optional[str] = None

class Device(BaseModel):
    id: int
    device_id: str
    name: Optional[str] = None
    fw_version: Optional[str] = None
    tags: List[str] = []
    type: Optional[str] = None
    location: Optional[str] = None
    last_seen: datetime

class DeviceCreate(BaseModel):
    device_id: str
    name: Optional[str] = None
    fw_version: Optional[str] = None
    tags: List[str] = []
    type: Optional[str] = None
    location: Optional[str] = None

class DeviceUpdate(BaseModel):
    name: Optional[str] = None
    fw_version: Optional[str] = None
    tags: Optional[List[str]] = None
    type: Optional[str] = None
    location: Optional[str] = None

class Sensors(BaseModel):
    temperature: Optional[float] = None
    humidity: Optional[float] = None
    distance: Optional[float] = None
    motion: Optional[bool] = None
    servo_state: Optional[str] = None
    led_states: Optional[dict] = None

class Net(BaseModel):
    tx_bytes: int = 0
    rx_bytes: int = 0
    connections: int = 0

class Telemetry(BaseModel):
    device_id: str
    ts: Optional[datetime] = None
    sensors: Optional[Sensors] = None
    net: Optional[Net] = None

class SuricataLog(BaseModel):
    event_ts: Optional[datetime] = None
    event_type: Optional[str] = None
    src_ip: Optional[str] = None
    src_port: Optional[str] = None
    dest_ip: Optional[str] = None
    dest_port: Optional[str] = None
    proto: Optional[str] = None
    signature: Optional[str] = None
    signature_id: Optional[str] = None
    severity: Optional[str] = None
    raw: Optional[dict] = None

# Import services
from .influxdb_data_service import InfluxDBDataService
from .ml_service import anomaly_service

# Initialize services
influx_data_service = None
influx_service = None  # Alias for backward compatibility

# Create FastAPI app
app = FastAPI(title="SIAC-IoT Backend", version="1.0.0")

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, specify your frontend URL
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup_event():
    """Initialize services on startup"""
    global influx_data_service, influx_service

    # Initialize InfluxDB service
    try:
        influx_data_service = InfluxDBDataService()
        influx_service = influx_data_service  # Alias for backward compatibility
        print("InfluxDB service initialized successfully")

        # Seed initial data
        influx_data_service.seed_initial_data()
        print("Initial data seeded successfully")
    except Exception as e:
        print(f"Failed to initialize InfluxDB service: {e}")
        influx_data_service = None
        influx_service = None

    # Initialize MQTT client
    init_mqtt_client()

# MQTT Client setup
mqtt_client = None
mqtt_connected = False

# WebSocket connections
websocket_connections = set()

def on_mqtt_connect(client, userdata, flags, rc):
    global mqtt_connected
    if rc == 0:
        mqtt_connected = True
        print("MQTT connected successfully")
        # Subscribe to ESP32 telemetry topics
        client.subscribe("devices/+/telemetry")
        print("Subscribed to devices/+/telemetry")
    else:
        mqtt_connected = False
        print(f"MQTT connection failed with code {rc}")

def on_mqtt_disconnect(client, userdata, rc):
    global mqtt_connected
    mqtt_connected = False
    print(f"MQTT disconnected with code {rc}")

def on_mqtt_message(client, userdata, msg):
    try:
        payload = json.loads(msg.payload.decode())
        device_id = msg.topic.split('/')[1]  # Extract device_id from topic
        
        # Process telemetry data
        process_telemetry(device_id, payload)
    except Exception as e:
        print(f"Error processing MQTT message: {e}")

def process_telemetry(device_id: str, payload: dict):
    """Process incoming telemetry from ESP32 devices"""
    try:
        # Extract sensor data
        sensors = payload.get('sensors', {})
        net = payload.get('net', {})
        
        # Write telemetry to InfluxDB
        if influx_data_service:
            influx_data_service.save_telemetry(
                device_id=device_id,
                temperature=sensors.get('temperature'),
                humidity=sensors.get('humidity'),
                distance=sensors.get('distance'),
                motion=sensors.get('motion'),
                servo_state=sensors.get('servo_state'),
                led_states=sensors.get('led_states'),
                tx_bytes=net.get('tx_bytes', 0),
                rx_bytes=net.get('rx_bytes', 0),
                connections=net.get('connections', 0),
                timestamp=datetime.utcnow()
            )
            
            # Update device last_seen
            influx_data_service.update_device_last_seen(device_id, datetime.utcnow())
            
            # Broadcast device status update
            import asyncio
            asyncio.create_task(broadcast_websocket_message({
                "type": "device_status",
                "device_id": device_id,
                "last_seen": datetime.utcnow().isoformat(),
                "status": "online"
            }))

        # Check for anomalies using ML service
        if anomaly_service:
            features = [
                sensors.get('temperature', 0) or 0,
                sensors.get('humidity', 0) or 0,
                sensors.get('distance', 0) or 0,
                1 if sensors.get('motion', False) else 0,
                net.get('tx_bytes', 0),
                net.get('rx_bytes', 0),
                net.get('connections', 0)
            ]
            
            is_anomaly, score = anomaly_service.predict_anomaly(features)
            
            if is_anomaly:
                # Create alert in InfluxDB
                if influx_data_service:
                    alert_id = str(uuid.uuid4())
                    influx_data_service.save_alert(
                        alert_id=alert_id,
                        device_id=device_id,
                        severity="high" if score > 0.8 else "medium",
                        score=score,
                        reason="Anomaly detected in sensor data",
                        acknowledged=False,
                        timestamp=datetime.utcnow()
                    )

                    # Send email alert if configured
                    alert_data = {
                        "alert_id": alert_id,
                        "device_id": device_id,
                        "severity": "high" if score > 0.8 else "medium",
                        "score": score,
                        "reason": "Anomaly detected in sensor data",
                        "ts": datetime.utcnow()
                    }
                    maybe_send_email_alert(alert_data)
                    
                    # Broadcast alert via WebSocket
                    import asyncio
                    asyncio.create_task(broadcast_websocket_message({
                        "type": "alert",
                        "alert_id": alert_id,
                        "device_id": device_id,
                        "severity": "high" if score > 0.8 else "medium",
                        "score": score,
                        "reason": "Anomaly detected in sensor data",
                        "ts": datetime.utcnow().isoformat()
                    }))
        
        # Broadcast telemetry update via WebSocket
        import asyncio
        asyncio.create_task(broadcast_websocket_message({
            "type": "telemetry",
            "device_id": device_id,
            "sensors": sensors,
            "net": net,
            "ts": datetime.utcnow().isoformat()
        }))
        
        print(f"Processed telemetry for device {device_id}")
        
    except Exception as e:
        print(f"Error processing telemetry: {e}")

def init_mqtt_client():
    global mqtt_client
    mqtt_host = os.environ.get("MQTT_HOST")
    mqtt_port = int(os.environ.get("MQTT_PORT"))
    mqtt_user = os.environ.get("MQTT_USERNAME")
    mqtt_pass = os.environ.get("MQTT_PASSWORD")
    mqtt_ca_cert = os.environ.get("MQTT_CA_CERT")
    
    mqtt_client = mqtt.Client()
    mqtt_client.on_connect = on_mqtt_connect
    mqtt_client.on_disconnect = on_mqtt_disconnect
    mqtt_client.on_message = on_mqtt_message
    
    if mqtt_user and mqtt_pass:
        mqtt_client.username_pw_set(mqtt_user, mqtt_pass)

    # TLS configuration
    mqtt_client.tls_set(ca_certs=mqtt_ca_cert, keyfile=None, tls_version=ssl.PROTOCOL_TLS_CLIENT)
    mqtt_client.tls_insecure_set(False)
    
    try:
        mqtt_client.connect(mqtt_host, mqtt_port, 60)
        mqtt_client.loop_start()  # Start in background thread
        print(f"MQTT client initialized for {mqtt_host}:{mqtt_port}")
    except Exception as e:
        print(f"Failed to initialize MQTT client: {e}")

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)

@app.get("/", include_in_schema=False)
def root():
    return {"status": "ok"}


@app.get("/api/v1/health")
def health():
    influx_status = "connected" if influx_service and influx_service.is_connected() else "disconnected"
    return {"status": "ok", "components": {"mqtt": "connected" if mqtt_connected else "disconnected", "influx": influx_status}}


async def broadcast_websocket_message(message: dict):
    """Broadcast message to all connected WebSocket clients"""
    disconnected = set()
    for websocket in websocket_connections:
        try:
            await websocket.send_json(message)
        except:
            disconnected.add(websocket)
    
    # Remove disconnected clients
    websocket_connections.difference_update(disconnected)


@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await websocket.accept()
    websocket_connections.add(websocket)

    try:
        # Keep connection alive for broadcasting, don't expect client messages
        while True:
            try:
                # Wait for any message with a timeout to keep connection alive
                await asyncio.wait_for(websocket.receive_text(), timeout=30.0)
                # If we get a message, echo it back for health check
                await websocket.send_json({"type": "pong", "timestamp": datetime.utcnow().isoformat()})
            except asyncio.TimeoutError:
                # Send a ping to keep connection alive
                await websocket.send_json({"type": "ping", "timestamp": datetime.utcnow().isoformat()})
    except Exception:
        # Connection closed or error
        pass
    finally:
        websocket_connections.discard(websocket)


@app.post("/api/v1/auth/login", response_model=AuthResponse)
def login(req: AuthRequest):
    if not influx_data_service:
        raise HTTPException(status_code=503, detail="Database service unavailable")

    user = influx_data_service.get_user(req.username)
    if not user or not influx_data_service.verify_password(req.password, user.get("password", "")):
        raise HTTPException(status_code=401, detail="Invalid credentials")

    return AuthResponse(
        access_token=str(uuid.uuid4()),
        refresh_token=str(uuid.uuid4()),
        username=user["username"],
        role=user.get("role", "admin"),
    )


# Users CRUD
@app.get("/api/v1/users", response_model=List[User])
def list_users():
    if not influx_data_service:
        raise HTTPException(status_code=503, detail="Database service unavailable")

    users = influx_data_service.list_users()
    # Convert to Pydantic User model format
    return [
        User(
            id=i+1,  # Mock ID for compatibility
            username=user["username"],
            email=user.get("email"),
            role=user.get("role", "admin"),
            is_active=user.get("is_active", True),
            created_at=datetime.utcnow()  # Mock created_at
        )
        for i, user in enumerate(users)
    ]


@app.post("/api/v1/users", response_model=User, status_code=201)
def create_user(payload: UserCreate):
    if not influx_data_service:
        raise HTTPException(status_code=503, detail="Database service unavailable")

    # Check if user already exists
    existing_user = influx_data_service.get_user(payload.username)
    if existing_user:
        raise HTTPException(status_code=409, detail="Username already exists")

    success = influx_data_service.create_user(
        username=payload.username,
        password=payload.password,
        email=payload.email,
        role=payload.role or "viewer"
    )

    if not success:
        raise HTTPException(status_code=500, detail="Failed to create user")

    # Return the created user
    user_data = influx_data_service.get_user(payload.username)
    return User(
        id=1,  # Mock ID
        username=user_data["username"],
        email=user_data.get("email"),
        role=user_data.get("role", "viewer"),
        is_active=user_data.get("is_active", True),
        created_at=datetime.utcnow()
    )


@app.put("/api/v1/users/{user_id}", response_model=User)
def update_user(user_id: int, payload: UserUpdate):
    if not influx_data_service:
        raise HTTPException(status_code=503, detail="Database service unavailable")

    # For now, we'll use username as identifier since InfluxDB uses tags
    # This is a limitation - we need to get user by some other means
    # For simplicity, let's assume user_id corresponds to username for now
    # In a real implementation, we'd need a proper user ID system

    # Get all users and find by index (temporary solution)
    users = influx_data_service.list_users()
    if user_id < 1 or user_id > len(users):
        raise HTTPException(status_code=404, detail="User not found")

    username = users[user_id - 1]["username"]

    update_data = {}
    if payload.email is not None:
        update_data["email"] = payload.email
    if payload.role is not None:
        update_data["role"] = payload.role
    if payload.is_active is not None:
        update_data["is_active"] = payload.is_active
    if payload.password:
        update_data["password"] = payload.password

    success = influx_data_service.update_user(username, update_data)
    if not success:
        raise HTTPException(status_code=500, detail="Failed to update user")

    # Return updated user
    user_data = influx_data_service.get_user(username)
    return User(
        id=user_id,
        username=user_data["username"],
        email=user_data.get("email"),
        role=user_data.get("role", "admin"),
        is_active=user_data.get("is_active", True),
        created_at=datetime.utcnow()
    )


@app.delete("/api/v1/users/{user_id}", status_code=204)
def delete_user(user_id: int):
    if not influx_data_service:
        raise HTTPException(status_code=503, detail="Database service unavailable")

    # Get all users and find by index (temporary solution)
    users = influx_data_service.list_users()
    if user_id < 1 or user_id > len(users):
        raise HTTPException(status_code=404, detail="User not found")

    username = users[user_id - 1]["username"]

    success = influx_data_service.delete_user(username)
    if not success:
        raise HTTPException(status_code=500, detail="Failed to delete user")

    return JSONResponse(status_code=204, content=None)


@app.get("/api/v1/devices", response_model=List[Device])
def list_devices():
    if not influx_data_service:
        raise HTTPException(status_code=503, detail="Database service unavailable")

    devices = influx_data_service.list_devices()
    # Convert to Pydantic Device model format
    return [
        Device(
            id=i+1,  # Mock ID for compatibility
            device_id=device["device_id"],
            name=device.get("name"),
            fw_version=device.get("fw_version"),
            tags=device.get("tags", []),
            type=device.get("type"),
            location=device.get("location"),
            last_seen=datetime.utcnow()  # Mock last_seen
        )
        for i, device in enumerate(devices)
    ]


@app.get("/api/v1/devices/{device_id}", response_model=Device)
def get_device(device_id: str):
    if not influx_data_service:
        raise HTTPException(status_code=503, detail="Database service unavailable")

    device = influx_data_service.get_device(device_id)
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")

    return Device(
        id=1,  # Mock ID
        device_id=device["device_id"],
        name=device.get("name"),
        fw_version=device.get("fw_version"),
        tags=device.get("tags", []),
        type=device.get("type"),
        location=device.get("location"),
        last_seen=datetime.utcnow()
    )


@app.post("/api/v1/devices", response_model=Device, status_code=201)
def create_device(payload: DeviceCreate):
    if not influx_data_service:
        raise HTTPException(status_code=503, detail="Database service unavailable")

    # Check if device already exists
    existing_device = influx_data_service.get_device(payload.device_id)
    if existing_device:
        raise HTTPException(status_code=409, detail="Device already exists")

    from .influxdb_data_service import DeviceData
    device_data = DeviceData(
        device_id=payload.device_id,
        name=payload.name,
        fw_version=payload.fw_version,
        tags=payload.tags,
        type=payload.type,
        location=payload.location
    )

    success = influx_data_service.create_device(device_data)
    if not success:
        raise HTTPException(status_code=500, detail="Failed to create device")

    # Return the created device
    device = influx_data_service.get_device(payload.device_id)
    return Device(
        id=1,  # Mock ID
        device_id=device["device_id"],
        name=device.get("name"),
        fw_version=device.get("fw_version"),
        tags=device.get("tags", []),
        type=device.get("type"),
        location=device.get("location"),
        last_seen=datetime.utcnow()
    )


@app.put("/api/v1/devices/{device_id}", response_model=Device)
def update_device(device_id: str, payload: DeviceUpdate):
    if not influx_data_service:
        raise HTTPException(status_code=503, detail="Database service unavailable")

    # Check if device exists
    existing_device = influx_data_service.get_device(device_id)
    if not existing_device:
        raise HTTPException(status_code=404, detail="Device not found")

    update_data = {}
    if payload.name is not None:
        update_data["name"] = payload.name
    if payload.fw_version is not None:
        update_data["fw_version"] = payload.fw_version
    if payload.tags is not None:
        update_data["tags"] = payload.tags
    if payload.type is not None:
        update_data["type"] = payload.type
    if payload.location is not None:
        update_data["location"] = payload.location

    success = influx_data_service.update_device(device_id, update_data)
    if not success:
        raise HTTPException(status_code=500, detail="Failed to update device")

    # Return updated device
    device = influx_data_service.get_device(device_id)
    return Device(
        id=1,  # Mock ID
        device_id=device["device_id"],
        name=device.get("name"),
        fw_version=device.get("fw_version"),
        tags=device.get("tags", []),
        type=device.get("type"),
        location=device.get("location"),
        last_seen=datetime.utcnow()
    )


@app.post("/api/v1/devices/{device_id}/commands")
def send_device_command(device_id: str, cmd_green: Optional[bool] = None, cmd_red: Optional[bool] = None):
    """Send command to device and record in InfluxDB"""
    if not mqtt_connected:
        raise HTTPException(status_code=503, detail="MQTT not connected")

    # Prepare command payload
    command_payload = {}
    if cmd_green is not None:
        command_payload["cmd_green"] = cmd_green
    if cmd_red is not None:
        command_payload["cmd_red"] = cmd_red

    if not command_payload:
        raise HTTPException(status_code=400, detail="No command specified")

    # Publish to MQTT
    try:
        mqtt_client.publish(f"devices/{device_id}/commands", json.dumps(command_payload))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to send MQTT command: {str(e)}")

    # Write command to InfluxDB
    if influx_service:
        influx_service.write_command(
            device_id=device_id,
            cmd_green=cmd_green,
            cmd_red=cmd_red,
            timestamp=datetime.utcnow()
        )

    return {"status": "command_sent", "device_id": device_id, "commands": command_payload}


@app.post("/api/v1/telemetry", status_code=202)
def ingest_telemetry(t: Telemetry):
    if not influx_data_service:
        raise HTTPException(status_code=503, detail="Database service unavailable")

    # Create telemetry data
    from .influxdb_data_service import TelemetryData
    telemetry_data = TelemetryData(
        device_id=t.device_id,
        ts=t.ts or datetime.utcnow(),
        temperature=t.sensors.temperature if t.sensors and t.sensors.temperature is not None else None,
        humidity=t.sensors.humidity if t.sensors and t.sensors.humidity is not None else None,
        distance=t.sensors.distance if t.sensors and t.sensors.distance is not None else None,
        motion=t.sensors.motion if t.sensors and t.sensors.motion is not None else None,
        servo_state=t.sensors.servo_state if t.sensors and t.sensors.servo_state is not None else None,
        led_states=t.sensors.led_states if t.sensors and t.sensors.led_states is not None else None,
        tx_bytes=t.net.tx_bytes if t.net and t.net.tx_bytes is not None else 0,
        rx_bytes=t.net.rx_bytes if t.net and t.net.rx_bytes is not None else 0,
        connections=t.net.connections if t.net and t.net.connections is not None else 0
    )

    success = influx_data_service.save_telemetry(telemetry_data)
    if not success:
        raise HTTPException(status_code=500, detail="Failed to save telemetry")

    # Try ML model prediction
    is_anomaly = False
    model_used = False
    model_status = getattr(anomaly_service, 'model_status', 'unavailable') if anomaly_service else 'unavailable'

    try:
        payload = {
            'temperature': telemetry_data.temperature,
            'humidity': telemetry_data.humidity,
            'distance': telemetry_data.distance,
            'motion': telemetry_data.motion,
            'servo_state': telemetry_data.servo_state,
            'led_states': telemetry_data.led_states,
            'tx_bytes': telemetry_data.tx_bytes,
            'rx_bytes': telemetry_data.rx_bytes,
            'connections': telemetry_data.connections,
            'ts': telemetry_data.ts.isoformat(),
        }

        # Try ML model prediction
        if anomaly_service:
            pred_is_anom, anom_score, status = anomaly_service.predict_anomaly(payload)
            if status == 'trained':
                model_used = True
        else:
            pred_is_anom, anom_score, status = False, 0.0, "unavailable"
            if pred_is_anom:
                is_anomaly = True
                from .influxdb_data_service import AlertData
                alert_data = AlertData(
                    alert_id=str(uuid.uuid4()),
                    device_id=t.device_id,
                    ts=telemetry_data.ts,
                    severity="high",
                    score=float(-anom_score) if isinstance(anom_score, (int, float)) else 0.0,
                    reason=f"Anomalie détectée par modèle ML (score={anom_score:.4f})",
                    acknowledged=False,
                    metadata={"metric": "ml", "model": "isolation_forest"}
                )
                influx_data_service.save_alert(alert_data)
                maybe_send_email_alert(alert_data)
    except Exception:
        # On any model error, don't block ingestion
        pass

    # Fallback to simple z-score detection on temperature when model not ready
    if not model_used and telemetry_data.temperature is not None:
        # For simplicity, we'll skip the z-score calculation for now
        # This would require querying recent telemetry data
        pass

    return JSONResponse(status_code=202, content={
        "received": True,
        "device_id": t.device_id,
        "is_anomaly": is_anomaly,
        "model_used": model_used,
        "model_status": model_status,
    })


@app.post("/api/v1/predict")
def predict(t: Telemetry):
    # Dummy prediction: random threshold logic can be replaced by real model
    return {"score": 0.1, "is_anomaly": False, "threshold": 0.7}


@app.get("/api/v1/dashboard_summary")
def dashboard_summary():
    """Get dashboard summary from InfluxDB"""
    if not influx_service or not influx_service.is_connected():
        raise HTTPException(status_code=503, detail="InfluxDB not connected")

    try:
        summary = influx_service.get_dashboard_summary()
        return summary
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get dashboard summary: {str(e)}")


@app.get("/api/v1/alerts/recent")
def recent_alerts(limit: int = 5):
    """Get recent alerts from InfluxDB"""
    if not influx_service or not influx_service.is_connected():
        raise HTTPException(status_code=503, detail="InfluxDB not connected")

    try:
        alerts = influx_service.get_recent_alerts(limit=limit)
        return alerts
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get recent alerts: {str(e)}")


@app.get("/api/v1/alerts/active")
def active_alerts():
    """Get active alerts from InfluxDB"""
    if not influx_service or not influx_service.is_connected():
        raise HTTPException(status_code=503, detail="InfluxDB not connected")

    try:
        alerts = influx_service.get_active_alerts()
        return alerts
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get active alerts: {str(e)}")


@app.post("/api/v1/alerts/{alert_id}/ack")
def acknowledge_alert(alert_id: str):
    """Acknowledge an alert (not implemented for InfluxDB - time-series data is immutable)"""
    # Note: InfluxDB doesn't support updates to existing data
    # In a production system, you might want to store alert state separately
    raise HTTPException(status_code=501, detail="Alert acknowledgment not implemented for InfluxDB backend")


@app.post("/api/v1/alerts/{alert_id}/resolve")
def resolve_alert(alert_id: str):
    """Resolve an alert (not implemented for InfluxDB - time-series data is immutable)"""
    # Note: InfluxDB doesn't support updates to existing data
    # In a production system, you might want to store alert state separately
    raise HTTPException(status_code=501, detail="Alert resolution not implemented for InfluxDB backend")


@app.get("/api/v1/alerts/recommendations")
def get_alert_recommendations():
    """Get ML-generated alert recommendations from anomaly analysis"""
    if not influx_service or not influx_service.is_connected():
        raise HTTPException(status_code=503, detail="InfluxDB not connected")
    
    if not anomaly_service:
        raise HTTPException(status_code=503, detail="ML service not available")

    try:
        active_alerts = influx_service.get_active_alerts()
        recommendations = []

        for alert in active_alerts[:10]:  # Limit to 10 most recent
            # Utiliser le service ML pour générer des recommandations intelligentes
            ml_recommendation = anomaly_service.generate_recommendations(
                alert=alert,
                telemetry_history=None  # TODO: Optionally fetch telemetry history for trend analysis
            )
            
            rec = {
                "alert_id": alert.get("alert_id"),
                "device_id": alert.get("device_id"),
                "severity": alert.get("severity"),
                "reason": alert.get("reason"),
                "ml_status": ml_recommendation.get("status"),
                "priority": ml_recommendation.get("priority", "unknown"),
                "urgency": ml_recommendation.get("urgency", "à déterminer"),
                "confidence": ml_recommendation.get("confidence", 0.0),
                "ml_score": ml_recommendation.get("ml_score", 0.0),
                "root_cause": ml_recommendation.get("root_cause_analysis", []),
                "actions": ml_recommendation.get("recommendations", []),
                "timestamp": ml_recommendation.get("timestamp")
            }

            # Try to get device info
            device = influx_service.get_device(alert.get("device_id"))
            if device:
                rec["device_name"] = device.get("name")
                rec["device_location"] = device.get("location")
                rec["device_type"] = device.get("type")

            recommendations.append(rec)

        return {
            "count": len(recommendations),
            "ml_enabled": anomaly_service.model_status == "trained",
            "recommendations": recommendations
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get alert recommendations: {str(e)}")


@app.get("/api/v1/ml/status")
def get_ml_status():
    """
    Retourne le statut du modèle ML.
    """
    try:
        if anomaly_service:
            status = anomaly_service.get_status()
            return status
        else:
            return {"status": "error", "message": "ML service not available"}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.post("/api/v1/ml/train")
def train_ml_model(n_samples: int = 1000):
    """
    Force l'entraînement du modèle ML (pour admin).
    """
    try:
        if not anomaly_service:
            raise HTTPException(status_code=503, detail="ML service not available")
        
        success = anomaly_service.train_on_simulated_data(n_samples=n_samples)
        if success:
            return {"status": "success", "message": f"Model trained with {n_samples} samples"}
        else:
            raise HTTPException(status_code=500, detail="Training failed")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# InfluxDB Endpoints
@app.get("/api/v1/influx/measurements/{measurement}")
def get_influx_measurements(measurement: str, limit: int = 20):
    """Get recent measurements from InfluxDB"""
    if not influx_service or not influx_service.is_connected():
        raise HTTPException(status_code=503, detail="InfluxDB not connected")

    try:
        result = influx_service.get_recent_measurements(measurement, limit)
        if result:
            # Convert to simple format
            data = []
            for table in result:
                for record in table.records:
                    data.append({
                        "time": record.get_time().isoformat(),
                        "measurement": record.get_measurement(),
                        "device": record.values.get("device"),
                        "fields": {k: v for k, v in record.values.items() if not k.startswith("_")}
                    })
            return {"data": data}
        return {"data": []}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to query InfluxDB: {str(e)}")


@app.get("/api/v1/influx/sensor-data")
def get_sensor_data(device_id: Optional[str] = None, limit: int = 50):
    """Get sensor data from InfluxDB"""
    if not influx_service or not influx_service.is_connected():
        raise HTTPException(status_code=503, detail="InfluxDB not connected")

    try:
        device_filter = f'|> filter(fn: (r) => r.device == "{device_id}")' if device_id else ""
        flux_query = f'''
        from(bucket: "bucket_iot")
        |> range(start: -24h)
        |> filter(fn: (r) => r._measurement == "measurement")
        {device_filter}
        |> limit(n: {limit})
        |> sort(columns: ["_time"], desc: true)
        '''

        result = influx_service.query_data(flux_query)
        if result:
            data = []
            for table in result:
                for record in table.records:
                    data.append({
                        "time": record.get_time().isoformat(),
                        "device": record.values.get("device"),
                        "temperature": record.values.get("temperature"),
                        "humidity": record.values.get("humidity"),
                        "distance": record.values.get("distance")
                    })
            return {"data": data}
        return {"data": []}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to query sensor data: {str(e)}")


@app.get("/api/v1/influx/commands")
def get_commands(device_id: Optional[str] = None, limit: int = 20):
    """Get command history from InfluxDB"""
    if not influx_service or not influx_service.is_connected():
        raise HTTPException(status_code=503, detail="InfluxDB not connected")

    try:
        device_filter = f'|> filter(fn: (r) => r.device == "{device_id}")' if device_id else ""
        flux_query = f'''
        from(bucket: "bucket_iot")
        |> range(start: -24h)
        |> filter(fn: (r) => r._measurement == "commands")
        {device_filter}
        |> limit(n: {limit})
        |> sort(columns: ["_time"], desc: true)
        '''

        result = influx_service.query_data(flux_query)
        if result:
            data = []
            for table in result:
                for record in table.records:
                    data.append({
                        "time": record.get_time().isoformat(),
                        "device": record.values.get("device"),
                        "cmd_green": record.values.get("cmd_green"),
                        "cmd_red": record.values.get("cmd_red")
                    })
            return {"data": data}
        return {"data": []}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to query commands: {str(e)}")


@app.get("/api/v1/influx/alerts")
def get_influx_alerts(device_id: Optional[str] = None, limit: int = 20):
    """Get alerts from InfluxDB"""
    if not influx_service or not influx_service.is_connected():
        raise HTTPException(status_code=503, detail="InfluxDB not connected")

    try:
        device_filter = f'|> filter(fn: (r) => r.device == "{device_id}")' if device_id else ""
        flux_query = f'''
        from(bucket: "bucket_iot")
        |> range(start: -24h)
        |> filter(fn: (r) => r._measurement == "alerts")
        {device_filter}
        |> limit(n: {limit})
        |> sort(columns: ["_time"], desc: true)
        '''

        result = influx_service.query_data(flux_query)
        if result:
            data = []
            for table in result:
                for record in table.records:
                    data.append({
                        "time": record.get_time().isoformat(),
                        "device": record.values.get("device"),
                        "temp": record.values.get("temp"),
                        "temp_high": record.values.get("tempHigh"),
                        "temp_low": record.values.get("tempLow"),
                        "hum": record.values.get("hum"),
                        "hum_high": record.values.get("humHigh"),
                        "distance": record.values.get("distance"),
                        "distance_alert": record.values.get("distanceAlert"),
                        "message": record.values.get("message")
                    })
            return {"data": data}
        return {"data": []}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to query alerts: {str(e)}")


@app.get("/api/v1/telemetry/recent")
def recent_telemetry(limit: int = 20):
    """Get recent telemetry from InfluxDB"""
    if not influx_service or not influx_service.is_connected():
        raise HTTPException(status_code=503, detail="InfluxDB not connected")

    try:
        telemetry_data = influx_service.get_recent_telemetry(limit=limit)
        result = []
        for item in telemetry_data:
            result.append({
                "device_id": item.get("device_id"),
                "ts": item.get("ts").isoformat() if item.get("ts") else None,
                "sensors": {
                    "temperature": item.get("temperature"),
                    "humidity": item.get("humidity"),
                    "distance": item.get("distance"),
                    "motion": None,  # Not stored in current InfluxDB schema
                    "servo_state": None,  # Not stored in current InfluxDB schema
                    "led_states": None,  # Not stored in current InfluxDB schema
                },
                "net": {
                    "tx_bytes": item.get("tx_bytes"),
                    "rx_bytes": item.get("rx_bytes"),
                    "connections": item.get("connections"),
                }
            })
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get recent telemetry: {str(e)}")


@app.get("/api/v1/metrics/devices_activity_24h")
def devices_activity_24h():
    """Get device activity metrics from InfluxDB"""
    if not influx_service or not influx_service.is_connected():
        raise HTTPException(status_code=503, detail="InfluxDB not connected")

    try:
        # Get telemetry data from last 24h
        flux_query = '''
        from(bucket: "bucket_iot")
        |> range(start: -24h)
        |> filter(fn: (r) => r._measurement == "telemetry")
        |> aggregateWindow(every: 1h, fn: count, createEmpty: false)
        |> group(columns: ["device_id", "_start"])
        |> sort(columns: ["_start"], desc: true)
        '''

        result = influx_service.query_data(flux_query)
        device_counts = {}
        alert_counts = {}

        # Process telemetry data
        if result:
            for table in result:
                for record in table.records:
                    start_time = record.get_start()
                    device_id = record.values.get("device_id")
                    count = record.get_value()

                    hour_key = start_time.replace(minute=0, second=0, microsecond=0)
                    if hour_key not in device_counts:
                        device_counts[hour_key] = set()
                    if device_id:
                        device_counts[hour_key].add(device_id)

        # Get alerts data from last 24h
        alert_query = '''
        from(bucket: "bucket_iot")
        |> range(start: -24h)
        |> filter(fn: (r) => r._measurement == "alerts")
        |> aggregateWindow(every: 1h, fn: count, createEmpty: false)
        |> sort(columns: ["_start"], desc: true)
        '''

        alert_result = influx_service.query_data(alert_query)
        if alert_result:
            for table in alert_result:
                for record in table.records:
                    start_time = record.get_start()
                    count = record.get_value()
                    hour_key = start_time.replace(minute=0, second=0, microsecond=0)
                    alert_counts[hour_key] = count

        # Build response for last 24h (6 points, every 4 hours)
        now = datetime.utcnow()
        out = []
        for i in range(24, -1, -4):
            t = now - timedelta(hours=i)
            hk = t.replace(minute=0, second=0, microsecond=0)
            time_label = hk.strftime("%H:00")
            devices = len(device_counts.get(hk, set()))
            alerts = alert_counts.get(hk, 0)
            out.append({
                "time": time_label,
                "devices": devices,
                "alerts": alerts,
            })
        return out
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get device activity: {str(e)}")


@app.get("/api/v1/metrics/data_volume_7d")
def data_volume_7d():
    """Get data volume metrics from InfluxDB"""
    if not influx_service or not influx_service.is_connected():
        raise HTTPException(status_code=503, detail="InfluxDB not connected")

    try:
        # Get data volume from last 7 days
        flux_query = '''
        from(bucket: "bucket_iot")
        |> range(start: -7d)
        |> filter(fn: (r) => r._measurement == "telemetry")
        |> filter(fn: (r) => r._field == "tx_bytes" or r._field == "rx_bytes")
        |> aggregateWindow(every: 1d, fn: sum, createEmpty: false)
        |> group(columns: ["_start"])
        |> sort(columns: ["_start"], desc: true)
        '''

        result = influx_service.query_data(flux_query)
        daily_volumes = {}

        if result:
            for table in result:
                for record in table.records:
                    start_time = record.get_start()
                    day_key = datetime(start_time.year, start_time.month, start_time.day)
                    volume = record.get_value() or 0
                    daily_volumes[day_key] = daily_volumes.get(day_key, 0) + volume

        # Build response for last 7 days
        now = datetime.utcnow()
        out = []
        for i in range(6, -1, -1):
            d = now - timedelta(days=i)
            dk = datetime(d.year, d.month, d.day)
            vol_gb = round((daily_volumes.get(dk, 0)) / (1024 ** 3), 2)
            out.append({
                "day": ["Lun","Mar","Mer","Jeu","Ven","Sam","Dim"][dk.weekday()],
                "volume": vol_gb,
            })
        return out
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get data volume: {str(e)}")


# InfluxDB Sensor Data Endpoint
@app.get("/api/v1/influx/sensor-data")
def get_influx_sensor_data(device_id: Optional[str] = None, limit: int = 50):
    """Get sensor data from InfluxDB for charting"""
    if not influx_service or not influx_service.is_connected():
        raise HTTPException(status_code=503, detail="InfluxDB not connected")

    try:
        # Get recent telemetry data and aggregate by time for charting
        device_filter = f'|> filter(fn: (r) => r.device_id == "{device_id}")' if device_id else ""

        flux_query = f'''
        from(bucket: "{influx_service.bucket}")
        |> range(start: -24h)
        |> filter(fn: (r) => r._measurement == "telemetry")
        {device_filter}
        |> sort(columns: ["_time"], desc: true)
        |> limit(n: {limit * 3})  // Get more records since we have multiple fields per timestamp
        '''

        result = influx_service.query_data(flux_query)

        # Aggregate data by timestamp
        time_series = {}

        if result:
            for table in result:
                for record in table.records:
                    ts = record["_time"]
                    time_key = ts.replace(second=0, microsecond=0)  # Round to minute

                    if time_key not in time_series:
                        time_series[time_key] = {
                            "time": time_key,
                            "temperature": None,
                            "humidity": None,
                            "distance": None
                        }

                    field = record["_field"]
                    value = record["_value"]

                    if field == "temperature":
                        time_series[time_key]["temperature"] = value
                    elif field == "humidity":
                        time_series[time_key]["humidity"] = value
                    elif field == "distance":
                        time_series[time_key]["distance"] = value

        # Convert to list and sort by time
        chart_data = list(time_series.values())
        chart_data.sort(key=lambda x: x["time"], reverse=True)

        # Limit the results
        chart_data = chart_data[:limit]

        # Format time for frontend
        for item in chart_data:
            item["time"] = item["time"].strftime("%H:%M")

        return chart_data

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get sensor data: {str(e)}")


# Suricata Security Logs Endpoints
@app.post("/api/v1/suricata/logs", status_code=202)
def ingest_suricata_log(log: SuricataLog):
    """Ingest Suricata log to InfluxDB"""
    if not influx_service or not influx_service.is_connected():
        raise HTTPException(status_code=503, detail="InfluxDB not connected")

    try:
        log_data = SuricataLogData(
            event_ts=log.event_ts,
            event_type=log.event_type,
            src_ip=log.src_ip,
            src_port=log.src_port,
            dest_ip=log.dest_ip,
            dest_port=log.dest_port,
            proto=log.proto,
            signature=log.signature,
            signature_id=log.signature_id,
            severity=log.severity,
            raw=log.raw,
        )
        success = influx_service.save_suricata_log(log_data)
        if success:
            return {"status": "created"}
        else:
            raise HTTPException(status_code=500, detail="Failed to save Suricata log")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to ingest Suricata log: {str(e)}")


@app.get("/api/v1/suricata/logs/recent")
def get_recent_suricata_logs(limit: int = 20):
    """Get recent Suricata logs from InfluxDB"""
    if not influx_service or not influx_service.is_connected():
        raise HTTPException(status_code=503, detail="InfluxDB not connected")

    try:
        logs = influx_service.get_recent_suricata_logs(limit=limit)
        return logs
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get recent Suricata logs: {str(e)}")


@app.get("/api/v1/suricata/logs/stats")
def get_suricata_stats():
    """Get Suricata log statistics from InfluxDB"""
    if not influx_service or not influx_service.is_connected():
        raise HTTPException(status_code=503, detail="InfluxDB not connected")

    try:
        # Get logs from last 24h
        flux_query = '''
        from(bucket: "bucket_iot")
        |> range(start: -24h)
        |> filter(fn: (r) => r._measurement == "suricata_alerts")
        |> sort(columns: ["_time"], desc: true)
        '''

        result = influx_service.query_data(flux_query)
        logs_24h = []
        total_logs = 0

        if result:
            for table in result:
                for record in table.records:
                    log_data = {
                        "signature": record["_value"] if record["_field"] == "signature" else None,
                        "severity": record["_value"] if record["_field"] == "severity" else None,
                    }
                    logs_24h.append(log_data)
                    total_logs += 1

        # Derive categories from signatures (simplified logic)
        def categorize_signature(signature):
            if not signature:
                return 'unknown'
            sig_lower = signature.lower()
            if 'mqtt' in sig_lower and 'tls' in sig_lower:
                return 'mqtt_no_tls'
            elif 'brute' in sig_lower or 'admin' in sig_lower:
                return 'brute_force'
            elif 'scan' in sig_lower or 'nmap' in sig_lower:
                return 'network_scan'
            elif 'dos' in sig_lower or 'flood' in sig_lower:
                return 'dos'
            elif 'tls' in sig_lower:
                return 'tls_error'
            elif 'docker' in sig_lower or '172.17' in sig_lower:
                return 'intrusion'
            else:
                return 'other'

        categories = {}
        severities = {}

        for log in logs_24h:
            cat = categorize_signature(log["signature"])
            categories[cat] = categories.get(cat, 0) + 1

            sev = log["severity"] or '3'
            severities[sev] = severities.get(sev, 0) + 1

        return {
            "total_logs": total_logs,
            "logs_24h": len(logs_24h),
            "categories": categories,
            "severities": severities,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get Suricata stats: {str(e)}")


@app.get("/api/v1/suricata/logs/alerts")
def get_suricata_alerts():
    """Get high-priority Suricata alerts from InfluxDB"""
    if not influx_service or not influx_service.is_connected():
        raise HTTPException(status_code=503, detail="InfluxDB not connected")

    try:
        # Get high-priority logs (severity '1' and '2')
        flux_query = '''
        from(bucket: "bucket_iot")
        |> range(start: -7d)
        |> filter(fn: (r) => r._measurement == "suricata_alerts")
        |> filter(fn: (r) => r._field == "severity" and (r._value == "1" or r._value == "2"))
        |> sort(columns: ["_time"], desc: true)
        |> limit(n: 50)
        '''

        result = influx_service.query_data(flux_query)
        alerts = []

        if result:
            for table in result:
                for record in table.records:
                    alert = {
                        "event_ts": record["_time"],
                        "severity": record["_value"] if record["_field"] == "severity" else None,
                    }
                    # Get other fields for this record
                    for field_record in table.records:
                        if field_record["_time"] == record["_time"]:
                            field = field_record["_field"]
                            value = field_record["_value"]
                            if field != "severity":
                                alert[field] = value
                    alerts.append(alert)

        return alerts
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get Suricata alerts: {str(e)}")


@app.get("/api/v1/suricata/logs/export")
async def export_suricata_logs(format: str = "excel"):
    """Export Suricata logs to Excel or PDF"""
    if not influx_service or not influx_service.is_connected():
        raise HTTPException(status_code=503, detail="InfluxDB not connected")

    try:
        # Get all recent Suricata logs
        logs = influx_service.get_recent_suricata_logs(limit=1000)
        
        if format.lower() == "excel":
            # Create Excel file
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Suricata Alerts"
            
            # Headers
            headers = ["Timestamp", "Signature", "Severity", "Source IP", "Source Port", 
                      "Destination IP", "Destination Port", "Protocol", "Action", "Category"]
            ws.append(headers)
            
            # Style headers
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Add data
            for log in logs:
                severity = log.get('severity', '')
                severity_label = 'CRITIQUE' if severity == 1 else 'ÉLEVÉ' if severity == 2 else 'MOYEN' if severity == 3 else 'FAIBLE'
                
                ws.append([
                    str(log.get('event_ts', '')),
                    log.get('signature', ''),
                    severity_label,
                    log.get('src_ip', ''),
                    log.get('src_port', ''),
                    log.get('dest_ip', ''),
                    log.get('dest_port', ''),
                    log.get('proto', ''),
                    log.get('action', ''),
                    log.get('category', '')
                ])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return Response(
                content=output.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=suricata_alerts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
            )
            
        elif format.lower() == "pdf":
            # Create PDF file
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            import io
            
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
            elements = []
            
            # Title
            styles = getSampleStyleSheet()
            title = Paragraph("<b>Rapport d'Alertes Suricata IDS</b>", styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 12))
            
            # Subtitle
            subtitle = Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M:%S')}", styles['Normal'])
            elements.append(subtitle)
            elements.append(Spacer(1, 20))
            
            # Table data
            data = [["Timestamp", "Signature", "Sév.", "IP Source", "IP Dest", "Action"]]
            
            for log in logs[:100]:  # Limit to 100 for PDF
                severity = log.get('severity', '')
                severity_label = 'CRIT' if severity == 1 else 'ÉLEV' if severity == 2 else 'MOY' if severity == 3 else 'FAIB'
                
                data.append([
                    str(log.get('event_ts', ''))[:19],
                    str(log.get('signature', ''))[:40],
                    severity_label,
                    log.get('src_ip', ''),
                    log.get('dest_ip', ''),
                    log.get('action', '')
                ])
            
            # Create table
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D32F2F')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
            doc.build(elements)
            
            buffer.seek(0)
            return Response(
                content=buffer.getvalue(),
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename=suricata_alerts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"}
            )
        
        else:
            raise HTTPException(status_code=400, detail="Format not supported. Use 'excel' or 'pdf'")
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export Suricata logs: {str(e)}")


# Export endpoints
@app.get("/api/v1/telemetry/export")
async def export_telemetry(format: str = "excel"):
    """Export telemetry data from InfluxDB"""
    if not influx_service or not influx_service.is_connected():
        raise HTTPException(status_code=503, detail="InfluxDB not connected")

    try:
        # Get all telemetry data
        telemetry_data = influx_service.get_recent_telemetry(limit=10000)  # Large limit for export

        data = []
        for item in telemetry_data:
            data.append({
                "device_id": item.get("device_id"),
                "ts": item.get("ts").isoformat() if item.get("ts") else None,
                "temperature": item.get("temperature"),
                "humidity": item.get("humidity"),
                "distance": item.get("distance"),
                "tx_bytes": item.get("tx_bytes"),
                "rx_bytes": item.get("rx_bytes"),
                "connections": item.get("connections")
            })

        if format == "excel":
            df = pd.DataFrame(data)
            buffer = io.BytesIO()
            df.to_excel(buffer, index=False, engine='openpyxl')
            buffer.seek(0)
            return FileResponse(buffer, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename='telemetry.xlsx')
        elif format == "pdf":
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            elements = []
            table_data = [["Device ID", "Timestamp", "Temp", "Humidity", "Distance", "TX", "RX", "Connections"]]
            for d in data:
                table_data.append([d["device_id"], str(d["ts"]), d["temperature"], d["humidity"], d["distance"], d["tx_bytes"], d["rx_bytes"], d["connections"]])
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)
            doc.build(elements)
            buffer.seek(0)
            return FileResponse(buffer, media_type='application/pdf', filename='telemetry.pdf')
        else:
            raise HTTPException(status_code=400, detail="Invalid format. Use 'excel' or 'pdf'")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export telemetry: {str(e)}")


@app.get("/api/v1/alerts/export")
async def export_alerts(format: str = "excel"):
    """Export alerts data from InfluxDB"""
    if not influx_service or not influx_service.is_connected():
        raise HTTPException(status_code=503, detail="InfluxDB not connected")

    try:
        # Get all alerts data
        alerts_data = influx_service.get_recent_alerts(limit=10000)  # Large limit for export

        data = []
        for item in alerts_data:
            data.append({
                "alert_id": item.get("alert_id"),
                "device_id": item.get("device_id"),
                "ts": item.get("ts").isoformat() if item.get("ts") else None,
                "severity": item.get("severity"),
                "score": item.get("score"),
                "reason": item.get("reason"),
                "acknowledged": item.get("acknowledged"),
            })

        if format == "excel":
            df = pd.DataFrame(data)
            buffer = io.BytesIO()
            df.to_excel(buffer, index=False, engine='openpyxl')
            buffer.seek(0)
            return FileResponse(buffer, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename='alerts.xlsx')
        elif format == "pdf":
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            elements = []
            table_data = [["Alert ID", "Device ID", "Timestamp", "Severity", "Score", "Reason", "Acknowledged"]]
            for d in data:
                table_data.append([d["alert_id"], d["device_id"], str(d["ts"]), d["severity"], d["score"], d["reason"], d["acknowledged"]])
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)
            doc.build(elements)
            buffer.seek(0)
            return FileResponse(buffer, media_type='application/pdf', filename='alerts.pdf')
        else:
            raise HTTPException(status_code=400, detail="Invalid format. Use 'excel' or 'pdf'")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export alerts: {str(e)}")


@app.get("/api/v1/logs/export")
async def export_logs(format: str = "excel"):
    """Export Suricata logs from InfluxDB"""
    if not influx_service or not influx_service.is_connected():
        raise HTTPException(status_code=503, detail="InfluxDB not connected")

    try:
        # Get all Suricata logs
        logs_data = influx_service.get_recent_suricata_logs(limit=10000)  # Large limit for export

        data = []
        for item in logs_data:
            data.append({
                "event_ts": item.get("event_ts").isoformat() if item.get("event_ts") else None,
                "event_type": item.get("event_type"),
                "src_ip": item.get("src_ip"),
                "src_port": item.get("src_port"),
                "dest_ip": item.get("dest_ip"),
                "dest_port": item.get("dest_port"),
                "proto": item.get("proto"),
                "signature": item.get("signature"),
                "signature_id": item.get("signature_id"),
                "severity": item.get("severity"),
            })

        if format == "excel":
            df = pd.DataFrame(data)
            buffer = io.BytesIO()
            df.to_excel(buffer, index=False, engine='openpyxl')
            buffer.seek(0)
            return FileResponse(buffer, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename='logs.xlsx')
        elif format == "pdf":
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            elements = []
            table_data = [["Event TS", "Event Type", "Src IP", "Src Port", "Dest IP", "Dest Port", "Proto", "Signature", "Sig ID", "Severity"]]
            for d in data:
                table_data.append([str(d["event_ts"]), d["event_type"], d["src_ip"], d["src_port"], d["dest_ip"], d["dest_port"], d["proto"], d["signature"], d["signature_id"], d["severity"]])
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)
            doc.build(elements)
            buffer.seek(0)
            return FileResponse(buffer, media_type='application/pdf', filename='logs.pdf')
        else:
            raise HTTPException(status_code=400, detail="Invalid format. Use 'excel' or 'pdf'")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export logs: {str(e)}")


def maybe_send_email_alert(alert):
    try:
        host = os.environ.get("SMTP_HOST")
        to_addr = os.environ.get("ALERT_EMAIL_TO")
        if not host or not to_addr:
            return
        port = int(os.environ.get("SMTP_PORT", "25"))
        user = os.environ.get("SMTP_USER")
        pwd = os.environ.get("SMTP_PASS")
        sender = os.environ.get("SMTP_FROM", user or "alerts@siac.local")
        subj = f"[SIAC-IoT] Alerte {alert.get('severity', 'unknown').upper()} - {alert.get('device_id', 'unknown')}"
        body = f"Device: {alert.get('device_id', 'unknown')}\nTime: {alert.get('ts', 'unknown')}\nReason: {alert.get('reason', 'unknown')}\nScore: {alert.get('score', 0)}"
        msg = MIMEText(body)
        msg['Subject'] = subj
        msg['From'] = sender
        msg['To'] = to_addr
        with smtplib.SMTP(host, port, timeout=5) as s:
            if user and pwd:
                s.starttls()
                s.login(user, pwd)
            s.sendmail(sender, [to_addr], msg.as_string())
    except Exception:
        # Best-effort only; ignore failures
        pass
